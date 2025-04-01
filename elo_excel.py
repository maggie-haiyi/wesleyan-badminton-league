import openpyxl
from openpyxl.styles import Font
import os

# ELO Constants
INITIAL_ELO = 1500
K_FACTOR = 32
FILE_NAME = "badminton_league.xlsx"

class BadmintonLeagueExcel:
    def __init__(self):
        if not os.path.exists(FILE_NAME):
            self.create_workbook()
        self.load_workbook()

    def create_workbook(self):
        wb = openpyxl.Workbook()
        ws_players = wb.active
        ws_players.title = "Players"
        ws_players.append(["Name", "ELO Rating", "Match History", "Match Count"])
        ws_players['A1'].font = ws_players['B1'].font = ws_players['C1'].font = ws_players['D1'].font = Font(bold=True)
        
        ws_matches = wb.create_sheet("Match History")
        ws_matches.append(["Player 1", "Score 1", "Player 2", "Score 2", "ELO 1 Before", "ELO 1 After", "ELO 2 Before", "ELO 2 After"])
        ws_matches['A1'].font = ws_matches['B1'].font = ws_matches['C1'].font = ws_matches['D1'].font = Font(bold=True)
        
        wb.save(FILE_NAME)

    def load_workbook(self):
        self.wb = openpyxl.load_workbook(FILE_NAME)
        self.ws_players = self.wb["Players"]
        self.ws_matches = self.wb["Match History"]

    def save_workbook(self):
        self.wb.save(FILE_NAME)

    def get_players(self):
        players = {}
        for row in self.ws_players.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                try:
                    rating = float(row[1]) if isinstance(row[1], (int, float)) else INITIAL_ELO
                except ValueError:
                    rating = INITIAL_ELO
                history = row[2] if row[2] else ""
                count = int(row[3]) if isinstance(row[3], (int, float)) else 0
                players[row[0]] = (rating, history, count)
        return players

    def add_player(self, name):
        players = self.get_players()
        if name in players:
            print(f"Player '{name}' already exists.")
        else:
            self.ws_players.append([name, INITIAL_ELO, "", 0])
            self.save_workbook()
            print(f"Player '{name}' added with ELO {INITIAL_ELO}.")

    def update_elo(self, player1, player2, score1, score2):
        players = self.get_players()
        if player1 not in players or player2 not in players:
            print("Both players must be registered.")
            return
        
        rating1, history1, count1 = players[player1]
        rating2, history2, count2 = players[player2]
        
        expected1 = 1 / (1 + 10 ** ((rating2 - rating1) / 400))
        expected2 = 1 / (1 + 10 ** ((rating1 - rating2) / 400))
        
        if score1 > score2:
            score1_result, score2_result = 1, 0
        elif score1 < score2:
            score1_result, score2_result = 0, 1
        else:
            score1_result, score2_result = 0.5, 0.5
        
        new_rating1 = rating1 + K_FACTOR * (score1_result - expected1)
        new_rating2 = rating2 + K_FACTOR * (score2_result - expected2)
        
        for row in self.ws_players.iter_rows(min_row=2):
            if row[0].value == player1:
                row[1].value = round(new_rating1, 2)
                row[2].value = f"{history1} vs {player2} ({score1}-{score2}) " if history1 else f"vs {player2} ({score1}-{score2})"
                row[3].value = count1 + 1
            if row[0].value == player2:
                row[1].value = round(new_rating2, 2)
                row[2].value = f"{history2} vs {player1} ({score2}-{score1}) " if history2 else f"vs {player1} ({score2}-{score1})"
                row[3].value = count2 + 1
        
        self.ws_matches.append([player1, score1, player2, score2, rating1, round(new_rating1, 2), rating2, round(new_rating2, 2)])
        self.sort_leaderboard()
        self.save_workbook()
        print(f"Updated ELO: {player1} ({new_rating1:.2f}), {player2} ({new_rating2:.2f})")

    def sort_leaderboard(self):
        players = self.get_players()
        sorted_players = sorted(players.items(), key=lambda x: x[1][0], reverse=True)
        
        self.ws_players.delete_rows(2, self.ws_players.max_row)
        for name, (rating, history, count) in sorted_players:
            self.ws_players.append([name, rating, history, count])

    def show_leaderboard(self):
        self.sort_leaderboard()
        self.save_workbook()
        print("Leaderboard updated in the spreadsheet.")

if __name__ == "__main__":
    league = BadmintonLeagueExcel()
    while True:
        print("\n1. Add Player\n2. Record Match\n3. Show Leaderboard\n4. Exit")
        choice = input("Choose an option: ")
        if choice == "1":
            name = input("Enter player name: ").strip()
            league.add_player(name)
        elif choice == "2":
            p1 = input("Enter Player 1 Name: ").strip()
            p2 = input("Enter Player 2 Name: ").strip()
            try:
                s1 = int(input(f"Enter {p1}'s Score: "))
                s2 = int(input(f"Enter {p2}'s Score: "))
                league.update_elo(p1, p2, s1, s2)
            except ValueError:
                print("Invalid input. Please enter numeric scores.")
        elif choice == "3":
            league.show_leaderboard()
        elif choice == "4":
            break
        else:
            print("Invalid choice. Please try again.")
