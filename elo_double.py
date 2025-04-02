import openpyxl
from openpyxl.styles import Font
import os
from itertools import permutations

# ELO Constants
INITIAL_ELO = 1500
K_FACTOR = 32
FILE_NAME = "badminton_doubles.xlsx"

class BadmintonLeagueDoubles:
    def __init__(self):
        if not os.path.exists(FILE_NAME):
            self.create_workbook()
        self.load_workbook()

    def create_workbook(self):
        wb = openpyxl.Workbook()
        ws_duos = wb.active
        ws_duos.title = "Duos"
        ws_duos.append(["Player 1", "Player 2", "ELO Rating", "Match History", "Match Count"])
        for cell in ws_duos[1]:
            cell.font = Font(bold=True)

        ws_matches = wb.create_sheet("Match History")
        ws_matches.append(["Duo 1", "Duo 2", "Score 1", "Score 2", "ELO 1 Before", "ELO 1 After", "ELO 2 Before", "ELO 2 After"])
        for cell in ws_matches[1]:
            cell.font = Font(bold=True)

        wb.save(FILE_NAME)

    def load_workbook(self):
        self.wb = openpyxl.load_workbook(FILE_NAME)
        self.ws_duos = self.wb["Duos"]
        self.ws_matches = self.wb["Match History"]

    def save_workbook(self):
        self.wb.save(FILE_NAME)

    def get_duos(self):
        duos = {}
        for row in self.ws_duos.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                players = tuple(sorted([row[0], row[1]]))
                rating = float(row[2]) if isinstance(row[2], (int, float)) else INITIAL_ELO
                history = row[3] if row[3] else ""
                count = int(row[4]) if isinstance(row[4], (int, float)) else 0
                duos[players] = (rating, history, count)
        return duos

    def create_duo(self, player1, player2):
        players = self.get_duos()
        duo = tuple(sorted([player1, player2]))
        if duo in players:
            print(f"Duo '{player1} & {player2}' already exists.")
        else:
            self.ws_duos.append([duo[0], duo[1], INITIAL_ELO, "", 0])
            self.save_workbook()
            print(f"Duo '{player1} & {player2}' created with ELO {INITIAL_ELO}.")

    def update_elo(self, team1, team2, score1, score2):
        duos = self.get_duos()
        team1 = tuple(sorted(team1))
        team2 = tuple(sorted(team2))

        if team1 not in duos or team2 not in duos:
            print("Both teams must be registered.")
            return

        rating1, history1, count1 = duos[team1]
        rating2, history2, count2 = duos[team2]

        expected1 = 1 / (1 + 10 ** ((rating2 - rating1) / 400))
        expected2 = 1 / (1 + 10 ** ((rating1 - rating2) / 400))

        if score1 > score2:
            score1_result, score2_result = 1, 0
        elif score1 < score2:
            score1_result, score2_result = 0, 1
        else:
            score1_result, score2_result = 0.5, 0.5

        new_rating1 = rating1 + K_FACTOR * (score1_result - expected1)
        new_rating2 = rating2 + K_FACTOR * (score2_result - expected2)

        for row in self.ws_duos.iter_rows(min_row=2):
            duo = tuple(sorted([row[0].value, row[1].value]))
            if duo == team1:
                row[2].value = round(new_rating1, 2)
                row[3].value = f"{history1} vs {team2} ({score1}-{score2}) " if history1 else f"vs {team2} ({score1}-{score2})"
                row[4].value = count1 + 1
            if duo == team2:
                row[2].value = round(new_rating2, 2)
                row[3].value = f"{history2} vs {team1} ({score2}-{score1}) " if history2 else f"vs {team1} ({score2}-{score1})"
                row[4].value = count2 + 1

        self.ws_matches.append([f"{team1}", f"{team2}", score1, score2, rating1, round(new_rating1, 2), rating2, round(new_rating2, 2)])
        self.sort_leaderboard()
        self.save_workbook()
        print(f"Updated ELO: {team1} ({new_rating1:.2f}), {team2} ({new_rating2:.2f})")

    def sort_leaderboard(self):
        duos = self.get_duos()
        sorted_duos = sorted(duos.items(), key=lambda x: x[1][0], reverse=True)

        self.ws_duos.delete_rows(2, self.ws_duos.max_row)
        for (player1, player2), (rating, history, count) in sorted_duos:
            self.ws_duos.append([player1, player2, rating, history, count])

    def show_leaderboard(self):
        self.sort_leaderboard()
        self.save_workbook()
        print("Leaderboard updated in the spreadsheet.")

if __name__ == "__main__":
    league = BadmintonLeagueDoubles()
    while True:
        print("\n1. Create New Duo\n2. Record Match\n3. Show Leaderboard\n4. Exit")
        choice = input("Choose an option: ")
        if choice == "1":
            p1 = input("Enter first player name: ").strip()
            p2 = input("Enter second player name: ").strip()
            league.create_duo(p1, p2)
        elif choice == "2":
            t1p1 = input("Enter Team 1 Player 1: ").strip()
            t1p2 = input("Enter Team 1 Player 2: ").strip()
            t2p1 = input("Enter Team 2 Player 1: ").strip()
            t2p2 = input("Enter Team 2 Player 2: ").strip()
            try:
                s1 = int(input("Enter Team 1's Score: "))
                s2 = int(input("Enter Team 2's Score: "))
                league.update_elo((t1p1, t1p2), (t2p1, t2p2), s1, s2)
            except ValueError:
                print("Invalid input. Please enter numeric scores.")
        elif choice == "3":
            league.show_leaderboard()
        elif choice == "4":
            break
        else:
            print("Invalid choice. Please try again.")

